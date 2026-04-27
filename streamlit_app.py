import csv
import io
import re
import zipfile
from pathlib import Path
from urllib.parse import urlparse, unquote
from concurrent.futures import ThreadPoolExecutor, as_completed

import requests
import streamlit as st

try:
    import openpyxl
except ImportError:
    openpyxl = None

st.set_page_config(page_title="Bulk Image Downloader", page_icon="📥", layout="wide")

USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0 Safari/537.36"
REQUEST_TIMEOUT = 60
MAX_WORKERS = 8
MAX_FILE_SIZE_MB = 50
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024

ALLOWED_IMAGE_CONTENT_TYPES = {
    "image/jpeg",
    "image/jpg",
    "image/png",
    "image/webp",
    "image/gif",
    "image/bmp",
    "image/tiff",
    "image/x-icon",
    "image/svg+xml",
}


def sanitize_filename(name: str) -> str:
    name = str(name).strip().replace("\x00", "")
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name or "downloaded_file"


def get_extension_from_content_type(content_type: str) -> str:
    if not content_type:
        return ""
    content_type = content_type.lower().split(";")[0].strip()
    mapping = {
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "image/bmp": ".bmp",
        "image/tiff": ".tif",
        "image/x-icon": ".ico",
        "image/svg+xml": ".svg",
    }
    return mapping.get(content_type, "")


def get_name_from_content_disposition(content_disposition: str) -> str:
    if not content_disposition:
        return ""

    match = re.search(r"filename\*=UTF-8''([^;]+)", content_disposition, flags=re.I)
    if match:
        return sanitize_filename(unquote(match.group(1).strip().strip('"')))

    match = re.search(r'filename="?([^";]+)"?', content_disposition, flags=re.I)
    if match:
        return sanitize_filename(unquote(match.group(1).strip()))

    return ""


def get_name_from_url(url: str) -> str:
    try:
        parsed = urlparse(url)
        raw_name = Path(unquote(parsed.path)).name
        return sanitize_filename(raw_name)
    except Exception:
        return ""


def ensure_extension(filename: str, content_type: str, url: str) -> str:
    current_ext = Path(filename).suffix
    if current_ext:
        return filename

    ext = get_extension_from_content_type(content_type)
    if ext:
        return filename + ext

    url_name = get_name_from_url(url)
    url_ext = Path(url_name).suffix
    if url_ext:
        return filename + url_ext

    return filename + ".jpg"


def make_unique_name(filename: str, used_names: set[str]) -> str:
    base = Path(filename).stem
    ext = Path(filename).suffix
    candidate = filename
    counter = 1
    while candidate.lower() in used_names:
        candidate = f"{base}_{counter}{ext}"
        counter += 1
    used_names.add(candidate.lower())
    return candidate


def build_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": USER_AGENT,
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            "Accept-Language": "en-US,en;q=0.9",
            "Connection": "keep-alive",
        }
    )
    return session


def normalize_content_type(content_type: str) -> str:
    return content_type.lower().split(";")[0].strip()


def looks_like_image_url(url: str) -> bool:
    name = get_name_from_url(url).lower()
    return Path(name).suffix in {
        ".jpg",
        ".jpeg",
        ".png",
        ".webp",
        ".gif",
        ".bmp",
        ".tif",
        ".tiff",
        ".ico",
        ".svg",
    }


def validate_image_response(response: requests.Response, original_url: str) -> None:
    content_type = normalize_content_type(response.headers.get("Content-Type", ""))

    if content_type and content_type not in ALLOWED_IMAGE_CONTENT_TYPES:
        if not looks_like_image_url(response.url) and not looks_like_image_url(original_url):
            raise ValueError(f"URL did not return an image. Content-Type was: {content_type}")

    content_length = response.headers.get("Content-Length", "")
    if content_length and content_length.isdigit():
        if int(content_length) > MAX_FILE_SIZE_BYTES:
            raise ValueError(f"File is too large. Limit is {MAX_FILE_SIZE_MB} MB.")


def read_response_bytes(response: requests.Response) -> bytes:
    content = io.BytesIO()
    downloaded = 0

    for chunk in response.iter_content(chunk_size=1024 * 64):
        if chunk:
            downloaded += len(chunk)
            if downloaded > MAX_FILE_SIZE_BYTES:
                raise ValueError(f"File is too large. Limit is {MAX_FILE_SIZE_MB} MB.")
            content.write(chunk)

    return content.getvalue()


def parse_urls_from_text(text: str) -> list[str]:
    urls = []
    for line in text.splitlines():
        value = line.strip()
        if value.startswith("http://") or value.startswith("https://"):
            urls.append(value)
    return dedupe_keep_order(urls)


def parse_urls_from_uploaded_file(uploaded_file) -> list[str]:
    raw = uploaded_file.getvalue()
    try:
        content = raw.decode("utf-8-sig")
    except Exception:
        content = raw.decode("latin-1")

    urls = []
    if uploaded_file.name.lower().endswith(".csv"):
        reader = csv.reader(io.StringIO(content))
        for row in reader:
            for cell in row:
                value = cell.strip()
                if value.startswith("http://") or value.startswith("https://"):
                    urls.append(value)
                    break
    else:
        urls.extend(parse_urls_from_text(content))

    return dedupe_keep_order(urls)


def parse_rename_csv(uploaded_file) -> list[dict]:
    raw = uploaded_file.getvalue()
    try:
        content = raw.decode("utf-8-sig")
    except Exception:
        content = raw.decode("latin-1")

    items = []
    reader = csv.reader(io.StringIO(content))

    for row in reader:
        if len(row) < 2:
            continue

        file_name = str(row[0]).strip()
        url = str(row[1]).strip()

        if not file_name or not url:
            continue

        if not (url.startswith("http://") or url.startswith("https://")):
            continue

        items.append(
            {
                "file_name": sanitize_filename(file_name),
                "url": url,
            }
        )

    return dedupe_rename_items_keep_order(items)


def get_excel_sheet_names(uploaded_file) -> list[str]:
    if openpyxl is None:
        st.error("Excel support needs openpyxl. Please install it using: pip install openpyxl")
        return []

    raw = uploaded_file.getvalue()
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def parse_rename_excel(uploaded_file, sheet_name: str) -> list[dict]:
    if openpyxl is None:
        st.error("Excel support needs openpyxl. Please install it using: pip install openpyxl")
        return []

    raw = uploaded_file.getvalue()
    workbook = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    try:
        sheet = workbook[sheet_name]

        items = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            if not row or len(row) < 2:
                continue

            file_name = "" if row[0] is None else str(row[0]).strip()
            url = "" if row[1] is None else str(row[1]).strip()

            if not file_name or not url:
                continue

            if not (url.startswith("http://") or url.startswith("https://")):
                continue

            items.append(
                {
                    "file_name": sanitize_filename(file_name),
                    "url": url,
                }
            )

        return dedupe_rename_items_keep_order(items)
    finally:
        workbook.close()


def parse_rename_file(uploaded_file, sheet_name: str = "") -> list[dict]:
    file_name = uploaded_file.name.lower()

    if file_name.endswith(".csv"):
        return parse_rename_csv(uploaded_file)

    if file_name.endswith(".xlsx") and sheet_name:
        return parse_rename_excel(uploaded_file, sheet_name)

    return []


def dedupe_keep_order(items: list[str]) -> list[str]:
    seen = set()
    result = []
    for item in items:
        if item not in seen:
            result.append(item)
            seen.add(item)
    return result


def dedupe_rename_items_keep_order(items: list[dict]) -> list[dict]:
    seen = set()
    result = []

    for item in items:
        key = item["url"]
        if key not in seen:
            result.append(item)
            seen.add(key)

    return result


def download_one(url: str, naming_mode: str, prefix: str, serial_number: int) -> dict:
    session = build_session()
    response = session.get(url, stream=True, timeout=REQUEST_TIMEOUT, allow_redirects=True)
    response.raise_for_status()
    validate_image_response(response, url)

    content_type = response.headers.get("Content-Type", "")
    content_disposition = response.headers.get("Content-Disposition", "")

    header_name = get_name_from_content_disposition(content_disposition)
    url_name = get_name_from_url(response.url) or get_name_from_url(url)

    if naming_mode == "Original name from server":
        chosen_name = header_name or url_name or f"downloaded_file_{serial_number}"
        name_source = "content-disposition" if header_name else "url"
    elif naming_mode == "CDN or URL name":
        chosen_name = url_name or header_name or f"downloaded_file_{serial_number}"
        name_source = "url" if url_name else "content-disposition"
    else:
        clean_prefix = sanitize_filename(prefix) or "image"
        chosen_name = f"{clean_prefix}_{serial_number}"
        name_source = "custom-prefix"

    chosen_name = sanitize_filename(chosen_name)
    chosen_name = ensure_extension(chosen_name, content_type, response.url)
    content_bytes = read_response_bytes(response)

    return {
        "url": url,
        "final_url": response.url,
        "status": "success",
        "file_name": chosen_name,
        "name_source": name_source,
        "content_type": content_type,
        "http_status": response.status_code,
        "error": "",
        "bytes": content_bytes,
    }


def download_one_with_rename(item: dict) -> dict:
    url = item["url"]
    requested_file_name = item["file_name"]

    session = build_session()
    response = session.get(url, stream=True, timeout=REQUEST_TIMEOUT, allow_redirects=True)
    response.raise_for_status()
    validate_image_response(response, url)

    content_type = response.headers.get("Content-Type", "")

    chosen_name = sanitize_filename(requested_file_name)
    chosen_name = ensure_extension(chosen_name, content_type, response.url)
    content_bytes = read_response_bytes(response)

    return {
        "url": url,
        "final_url": response.url,
        "status": "success",
        "file_name": chosen_name,
        "name_source": "uploaded-file-column-a",
        "content_type": content_type,
        "http_status": response.status_code,
        "error": "",
        "bytes": content_bytes,
    }


def build_zip_and_report(results: list[dict]) -> tuple[bytes, str]:
    used_names = set()
    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        report_rows = []

        for row in results:
            row_for_csv = {k: v for k, v in row.items() if k != "bytes"}

            if row["status"] == "success":
                unique_name = make_unique_name(row["file_name"], used_names)
                row_for_csv["file_name"] = unique_name
                zf.writestr(unique_name, row["bytes"])

            report_rows.append(row_for_csv)

        report_buffer = io.StringIO()
        writer = csv.DictWriter(
            report_buffer,
            fieldnames=[
                "url",
                "final_url",
                "status",
                "file_name",
                "name_source",
                "content_type",
                "http_status",
                "error",
            ],
        )
        writer.writeheader()
        writer.writerows(report_rows)

        zf.writestr("download_report.csv", report_buffer.getvalue().encode("utf-8-sig"))

    zip_buffer.seek(0)
    return zip_buffer.getvalue(), "bulk_images_download.zip"


def run_bulk_download(urls: list[str], naming_mode: str, prefix: str) -> list[dict]:
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {
            executor.submit(download_task_wrapper, url, naming_mode, prefix, index): url
            for index, url in enumerate(urls, start=1)
        }
        progress = st.progress(0)
        status = st.empty()

        completed = 0
        total = len(futures)

        for future in as_completed(futures):
            result = future.result()
            results.append(result)
            completed += 1
            progress.progress(completed / total)
            status.info(f"Processed {completed} of {total}")

        status.success(f"Completed {completed} of {total}")
    return results


def run_bulk_download_with_rename(items: list[dict]) -> list[dict]:
    results = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(download_rename_task_wrapper, item): item for item in items}
        progress = st.progress(0)
        status = st.empty()

        completed = 0
        total = len(futures)

        for future in as_completed(futures):
            result = future.result()
            results.append(result)
            completed += 1
            progress.progress(completed / total)
            status.info(f"Processed {completed} of {total}")

        status.success(f"Completed {completed} of {total}")
    return results


def download_task_wrapper(url: str, naming_mode: str, prefix: str, serial_number: int) -> dict:
    try:
        return download_one(url, naming_mode, prefix, serial_number)
    except Exception as e:
        return {
            "url": url,
            "final_url": "",
            "status": "failed",
            "file_name": "",
            "name_source": "",
            "content_type": "",
            "http_status": "",
            "error": str(e),
            "bytes": b"",
        }


def download_rename_task_wrapper(item: dict) -> dict:
    try:
        return download_one_with_rename(item)
    except Exception as e:
        return {
            "url": item.get("url", ""),
            "final_url": "",
            "status": "failed",
            "file_name": item.get("file_name", ""),
            "name_source": "uploaded-file-column-a",
            "content_type": "",
            "http_status": "",
            "error": str(e),
            "bytes": b"",
        }


st.title("📥 Bulk Image Downloader")
st.caption("Download images in bulk from URLs and save them in a ZIP with the best filename available.")

download_type = st.radio(
    "Select Download Type",
    ["Normal Bulk Download", "Bulk Download by Renaming"],
    horizontal=True,
)

if download_type == "Normal Bulk Download":
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Paste URLs")
        url_text = st.text_area(
            "One image URL per line",
            height=220,
            placeholder="https://example.com/image1.jpg\nhttps://example.com/image2.jpg",
        )

    with col2:
        st.subheader("Or upload TXT / CSV")
        uploaded_file = st.file_uploader("Upload a TXT or CSV file", type=["txt", "csv"])

    st.subheader("Download Options")
    opt1, opt2 = st.columns([2, 1])

    with opt1:
        naming_mode = st.selectbox(
            "Naming mode",
            ["Original name from server", "CDN or URL name", "Custom prefix + serial handled automatically"],
            index=0,
        )

    with opt2:
        prefix = st.text_input("Custom prefix", value="image")

    urls = []
    if url_text.strip():
        urls.extend(parse_urls_from_text(url_text))
    if uploaded_file is not None:
        urls.extend(parse_urls_from_uploaded_file(uploaded_file))
    urls = dedupe_keep_order(urls)

    st.write(f"Total valid URLs found: **{len(urls)}**")
    st.caption(f"Safety limit: each image must be {MAX_FILE_SIZE_MB} MB or less.")

    if st.button("Start Bulk Download", type="primary", width="stretch"):
        if not urls:
            st.error("Please paste URLs or upload a file first.")
        else:
            with st.spinner("Downloading images and preparing ZIP..."):
                results = run_bulk_download(urls, naming_mode, prefix)

            success_count = sum(1 for r in results if r["status"] == "success")
            failed_count = sum(1 for r in results if r["status"] == "failed")

            st.success(f"Done. Success: {success_count} | Failed: {failed_count}")

            preview_rows = []
            for row in results:
                preview_rows.append(
                    {
                        "status": row["status"],
                        "file_name": row["file_name"],
                        "name_source": row["name_source"],
                        "http_status": row["http_status"],
                        "url": row["url"],
                        "error": row["error"],
                    }
                )

            st.dataframe(preview_rows, width="stretch")

            zip_bytes, zip_name = build_zip_and_report(results)
            st.download_button(
                label="Download ZIP",
                data=zip_bytes,
                file_name=zip_name,
                mime="application/zip",
                width="stretch",
            )

else:
    st.subheader("Bulk Download by Renaming")
    st.info("Upload a CSV or Excel file where Column A has the file name and Column B has the image URL.")

    rename_uploaded_file = st.file_uploader(
        "Upload CSV or Excel file for renaming",
        type=["csv", "xlsx"],
        help="Column A = File Name, Column B = Image URL",
    )

    rename_items = []
    selected_sheet_name = ""

    if rename_uploaded_file is not None:
        if rename_uploaded_file.name.lower().endswith(".xlsx"):
            sheet_names = get_excel_sheet_names(rename_uploaded_file)

            if sheet_names:
                selected_sheet_name = st.selectbox(
                    "Select Excel Sheet",
                    sheet_names,
                    index=0,
                    help="Choose the sheet/tab that has Column A as file name and Column B as image URL.",
                )
                rename_items = parse_rename_file(rename_uploaded_file, selected_sheet_name)
        else:
            rename_items = parse_rename_file(rename_uploaded_file)

    st.write(f"Total valid rename rows found: **{len(rename_items)}**")
    st.caption(f"Safety limit: each image must be {MAX_FILE_SIZE_MB} MB or less.")

    if rename_items:
        preview_rows = []
        for item in rename_items:
            preview_rows.append(
                {
                    "file_name": item["file_name"],
                    "url": item["url"],
                }
            )
        st.dataframe(preview_rows, width="stretch")

    if st.button("Start Bulk Download by Renaming", type="primary", width="stretch"):
        if not rename_items:
            st.error("Please upload a valid CSV or Excel file first. Column A should have file name and Column B should have image URL.")
        else:
            with st.spinner("Downloading images with renamed file names and preparing ZIP..."):
                results = run_bulk_download_with_rename(rename_items)

            success_count = sum(1 for r in results if r["status"] == "success")
            failed_count = sum(1 for r in results if r["status"] == "failed")

            st.success(f"Done. Success: {success_count} | Failed: {failed_count}")

            preview_rows = []
            for row in results:
                preview_rows.append(
                    {
                        "status": row["status"],
                        "file_name": row["file_name"],
                        "name_source": row["name_source"],
                        "http_status": row["http_status"],
                        "url": row["url"],
                        "error": row["error"],
                    }
                )

            st.dataframe(preview_rows, width="stretch")

            zip_bytes, zip_name = build_zip_and_report(results)
            st.download_button(
                label="Download ZIP",
                data=zip_bytes,
                file_name=zip_name,
                mime="application/zip",
                width="stretch",
            )

st.markdown(
    """
    <hr style="margin-top:30px; margin-bottom:10px;">
    <div style="text-align:center; color:gray; font-size:14px;">
        © Designed and Developed by Pratik Adsare
    </div>
    """,
    unsafe_allow_html=True,
)
